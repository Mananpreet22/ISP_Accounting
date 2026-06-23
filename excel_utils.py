from openpyxl import Workbook
from openpyxl.styles import Font
from tkinter import filedialog, messagebox
import sqlite3
from datetime import datetime

DB_NAME = "isp_billing.db"


# =========================
# EXPORT TRANSACTIONS (RAW)
# =========================
def export_transactions_to_excel():

    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()

    cursor.execute("""
    SELECT 
        t.id,
        t.type,
        t.category,
        t.amount,
        t.note,
        t.date,
        t.bill_id,
        t.payment_date,
        t.created_at,
        t.customer,
        t.payment_method,
        b.receipt_no
    FROM transactions t
    LEFT JOIN bills b ON t.bill_id = b.id
""")
    rows = cursor.fetchall()
    columns = [desc[0] for desc in cursor.description]

    conn.close()

    if not rows:
        messagebox.showinfo("Info", "No transactions to export.")
        return

    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        title="Save Excel File"
    )

    if not file_path:
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    ws.append(columns)

    for row in rows:
        ws.append(row)

    wb.save(file_path)

    messagebox.showinfo("Success", "Excel exported successfully!")


# =========================
# MONTHLY FINANCIAL REPORT
# =========================
def export_monthly_report():

    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()

    # Current month format YYYY-MM
    month = datetime.now().strftime("%Y-%m")

    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        initialfile=f"ISP_Report_{month}.xlsx",
        filetypes=[("Excel files", "*.xlsx")]
    )

    if not file_path:
        return

    wb = Workbook()
    header_font = Font(bold=True)

    # =========================
    # SUMMARY SHEET
    # =========================
    ws = wb.active
    ws.title = "Summary"

    cursor.execute("SELECT COUNT(*) FROM customers")
    total_customers = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM bills WHERE status='PAID'")
    paid_bills = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM bills WHERE status!='PAID'")
    unpaid_bills = cursor.fetchone()[0]

    # ✅ Revenue using payment_date
    cursor.execute("""
        SELECT SUM(amount) FROM bills
        WHERE status='PAID' AND payment_date LIKE ?
    """, (f"{month}%",))
    revenue = cursor.fetchone()[0] or 0

    # ✅ Expenses using expense_date
    cursor.execute("""
        SELECT SUM(amount) FROM expenses
        WHERE expense_date LIKE ?
    """, (f"{month}%",))
    expenses = cursor.fetchone()[0] or 0

    profit = revenue - expenses
    arpu = revenue / total_customers if total_customers else 0

    ws.append(["Metric", "Value"])
    ws["A1"].font = header_font
    ws["B1"].font = header_font

    data = [
        ("Total Customers", total_customers),
        ("Paid Bills", paid_bills),
        ("Unpaid Bills", unpaid_bills),
        ("Revenue (This Month)", revenue),
        ("Expenses (This Month)", expenses),
        ("Net Profit", profit),
        ("ARPU", round(arpu, 2)),
    ]

    for row in data:
        ws.append(row)

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 20

    # =========================
    # INCOME SHEET (JOINED)
    # =========================
    ws_income = wb.create_sheet("Income")

    headers = ["Bill ID", "Customer Name", "Bill Month", "Amount", "Payment Date"]
    ws_income.append(headers)

    for col in range(1, len(headers) + 1):
        ws_income.cell(row=1, column=col).font = header_font

    cursor.execute("""
        SELECT b.id, c.name, b.bill_month, b.amount, b.payment_date
        FROM bills b
        JOIN customers c ON b.customer_id = c.id
        WHERE b.status='PAID' AND b.payment_date LIKE ?
    """, (f"{month}%",))

    total_income = 0

    for row in cursor.fetchall():
        ws_income.append(row)
        try:
            total_income += float(row[3])
        except:
            pass

    ws_income.append(["", "", "", "", ""])
    ws_income.append(["", "", "TOTAL", total_income, ""])

    # =========================
    # EXPENSE SHEET
    # =========================
    ws_expense = wb.create_sheet("Expenses")

    headers = ["ID", "Category", "Description", "Amount", "Payment Method", "Expense Date"]
    ws_expense.append(headers)

    for col in range(1, len(headers) + 1):
        ws_expense.cell(row=1, column=col).font = header_font

    cursor.execute("""
        SELECT id, category, description, amount, payment_method, expense_date
        FROM expenses
        WHERE expense_date LIKE ?
    """, (f"{month}%",))

    total_expense = 0

    for row in cursor.fetchall():
        ws_expense.append(row)
        try:
            total_expense += float(row[3])
        except:
            pass

    ws_expense.append(["", "", "", "", "", ""])
    ws_expense.append(["", "", "TOTAL", total_expense, "", ""])

    # =========================
    # AUTO COLUMN WIDTH
    # =========================
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_length = 0
            col_letter = col[0].column_letter

            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            sheet.column_dimensions[col_letter].width = max_length + 3

    wb.save(file_path)
    conn.close()

    messagebox.showinfo("Success", "Professional Monthly Report Exported!")